import uuid
from datetime import datetime

import pandas as pd

from models.station import DetectionResult, Report, FlagLevel
from services.detection import DetectorRegistry
from services.station_loader import RegionData


def _flag_level_order(level: FlagLevel) -> int:
    order = {FlagLevel.SEVERE: 3, FlagLevel.WARNING: 2, FlagLevel.GENERAL: 1, FlagLevel.INFO: 0}
    return order.get(level, 0)


def generate_report(
    station_name: str,
    results: list[DetectionResult],
) -> Report:
    counts = {"Severe": 0, "Warning": 0, "General": 0, "Info": 0}
    for r in results:
        key = r.flag_level.value
        counts[key] = counts.get(key, 0) + 1

    return Report(
        id=str(uuid.uuid4()),
        station_name=station_name,
        created_at=datetime.now(),
        total_flags=len(results),
        severe_count=counts["Severe"],
        warning_count=counts["Warning"],
        general_count=counts["General"],
        info_count=counts["Info"],
        status="completed",
    )


def export_excel(report: Report, results: list[DetectionResult], output_path: str) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "汇总统计"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF")
    severe_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    warning_fill = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")
    general_fill = PatternFill(start_color="6BCB77", end_color="6BCB77", fill_type="solid")
    info_fill = PatternFill(start_color="4D96FF", end_color="4D96FF", fill_type="solid")

    ws_summary.append(["站点", report.station_name])
    ws_summary.append(["检测时间", str(report.created_at)])
    ws_summary.append(["报告ID", report.id])
    ws_summary.append([])
    ws_summary.append(["标记级别", "数量"])
    ws_summary.append(["严重", report.severe_count])
    ws_summary.append(["警告", report.warning_count])
    ws_summary.append(["一般", report.general_count])
    ws_summary.append(["提示", report.info_count])
    ws_summary.append(["总计", report.total_flags])
    ws_summary.append([])

    detector_stats = {}
    for r in results:
        key = r.detector
        if key not in detector_stats:
            detector_stats[key] = {"Severe": 0, "Warning": 0, "General": 0, "Info": 0}
        detector_stats[key][r.flag_level.value] += 1

    ws_summary.append(["检测器 \\ 标记级别", "严重", "警告", "一般", "提示", "合计"])
    for detector, counts in detector_stats.items():
        total = sum(counts.values())
        ws_summary.append([detector, counts["Severe"], counts["Warning"], counts["General"], counts["Info"], total])

    columns = ["时间", "实测值", "期望值", "偏差", "检测器", "触发规则", "标记级别", "补充说明"]
    by_data_type = {}
    for r in results:
        by_data_type.setdefault(r.data_type, []).append(r)

    sheet_names = {"excerpt": "降雨摘录表", "daily": "逐日降水表", "monthly": "月年降水对照表", "period_max": "各时段最大降水量"}

    for data_type, sheet_name in sheet_names.items():
        items = by_data_type.get(data_type, [])
        if not items:
            continue

        ws = wb.create_sheet(title=sheet_name)
        ws.append(columns)
        for cell in ws[1]:
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for r in items:
            row_idx = ws.max_row + 1
            ws.append([
                r.datetime,
                r.value,
                r.expected_value if r.expected_value is not None else "",
                r.deviation if r.deviation is not None else "",
                r.detector,
                r.trigger_rule,
                r.flag_level.value,
                r.detail,
            ])
            flag_cell = ws.cell(row=row_idx, column=7)
            if r.flag_level == FlagLevel.SEVERE:
                flag_cell.fill = severe_fill
            elif r.flag_level == FlagLevel.WARNING:
                flag_cell.fill = warning_fill
            elif r.flag_level == FlagLevel.GENERAL:
                flag_cell.fill = general_fill
            elif r.flag_level == FlagLevel.INFO:
                flag_cell.fill = info_fill

    wb.save(output_path)


def run_pipeline(
    target_station_id: str,
    regions: list[RegionData],
    config: dict,
) -> tuple[Report, list[DetectionResult]]:
    all_results: list[DetectionResult] = []
    all_station_meta = []
    seen = set()
    for region in regions:
        for s in region.stations:
            key = s.get("name", "")
            if key and key not in seen:
                seen.add(key)
                all_station_meta.append(s)

    table_data_map: dict[str, list[pd.DataFrame]] = {}
    for region in regions:
        for table_type, df in region.tables.items():
            if df is not None and len(df) > 0:
                table_data_map.setdefault(table_type, []).append(df)

    all_detectors = DetectorRegistry.get_all()

    for detector in all_detectors:
        data_type = detector.data_type
        dfs = table_data_map.get(data_type, [])

        if not dfs:
            continue

        combined = pd.concat(dfs, ignore_index=True)
        combined = combined.drop_duplicates()
        if len(combined) == 0:
            continue

        try:
            det_results = detector.detect(combined, all_station_meta, config)
        except Exception as e:
            print(f"  [WARN] {detector.name} failed: {e}")
            continue

        for r in det_results:
            if r.station_id == target_station_id:
                all_results.append(r)

    report = generate_report(target_station_id, all_results)
    for r in all_results:
        r.report_id = report.id

    return report, all_results
