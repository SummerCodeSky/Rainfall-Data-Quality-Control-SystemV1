import io
import os
import tempfile
import zipfile
from pathlib import Path

from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import FileResponse, StreamingResponse

from services.station_loader import load_all_regions, deduplicate_stations
from services.pipeline import run_pipeline, export_excel
from services.database import (
    save_report,
    list_reports,
    get_report,
    get_results as db_get_results,
    query_reports,
    query_reports_grouped,
    delete_report,
    delete_reports_by_ids,
    delete_reports_by_batch,
    get_all_station_names,
    get_all_region_names,
    get_reports_by_batch,
    get_batch_detection_results,
)
from api.config import _load_config

DATA_DIR = os.environ.get("DATA_DIR", os.path.join(os.path.dirname(__file__), "../../data"))

router = APIRouter(prefix="/api", tags=["detection"])


@router.post("/detect")
async def run_detection(body: dict):
    station_name = body.get("station", "")
    if not station_name:
        raise HTTPException(400, "station is required")

    regions = load_all_regions(DATA_DIR)
    if not regions:
        raise HTTPException(404, "No regions found")

    all_stations = deduplicate_stations(regions)
    found = any(s.get("name") == station_name for s in all_stations)
    region_name = next((s.get("region", "") for s in all_stations if s.get("name") == station_name), "")

    if not found:
        raise HTTPException(404, f"Station {station_name} not found")

    config = _load_config()
    report, results = run_pipeline(station_name, regions, config)

    report_dict = report.model_dump()
    report_dict["created_at"] = report_dict["created_at"].isoformat()
    report_dict["region_name"] = region_name
    report_dict["results"] = [r.model_dump() for r in results]

    save_report(
        {"id": report.id, "station_name": station_name, "region_name": region_name,
         "created_at": report_dict["created_at"], "total_flags": report.total_flags,
         "severe_count": report.severe_count, "warning_count": report.warning_count,
         "general_count": report.general_count, "info_count": report.info_count, "status": "completed"},
        [r.model_dump() for r in results],
    )

    return report_dict


@router.get("/results")
async def get_results(
    report_id: str = Query(None),
    station: str = Query(None),
    flag_level: str = Query(None),
    detector: str = Query(None),
):
    rows = db_get_results(
        report_id=report_id,
        station=station,
        flag_level=flag_level,
        detector=detector,
    )
    for r in rows:
        r.setdefault("report_id", report_id or "")
    return rows


@router.get("/reports")
async def get_reports(
    view_type: str = Query("detail"),
    regions: str = Query(None),
    stations: str = Query(None),
    start_time: str = Query(None),
    end_time: str = Query(None),
    risk_filter: str = Query(None),
    batch_detect_time: str = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(None),
):
    region_list = [r.strip() for r in regions.split(",") if r.strip()] if regions else None
    station_list = [s.strip() for s in stations.split(",") if s.strip()] if stations else None

    if view_type == "summary":
        ps = page_size or 10
        batches, total = query_reports_grouped(
            regions=region_list,
            stations=station_list,
            start_time=start_time,
            end_time=end_time,
            risk_filter=risk_filter,
            batch_detect_time=batch_detect_time,
            page=page,
            page_size=ps,
        )
        return {
            "view_type": "summary",
            "batches": batches,
            "total": total,
            "page": page,
            "page_size": ps,
        }
    else:
        ps = page_size or 20
        records, total = query_reports(
            regions=region_list,
            stations=station_list,
            start_time=start_time,
            end_time=end_time,
            risk_filter=risk_filter,
            batch_detect_time=batch_detect_time,
            page=page,
            page_size=ps,
        )
        return {
            "view_type": "detail",
            "records": records,
            "total": total,
            "page": page,
            "page_size": ps,
        }


@router.get("/reports/stations")
async def get_stations():
    return get_all_station_names()


@router.get("/reports/regions")
async def get_regions():
    return get_all_region_names()


@router.delete("/reports/{report_id}")
async def delete_single_report(report_id: str):
    deleted = delete_report(report_id)
    if not deleted:
        raise HTTPException(404, "Report not found")
    return {"ok": True}


@router.post("/reports/batch-delete")
async def batch_delete_reports(body: dict):
    ids = body.get("ids")
    region = body.get("region")
    detect_time = body.get("detect_time")

    if ids:
        count = delete_reports_by_ids(ids)
        return {"ok": True, "deleted": count}
    elif region and detect_time:
        count = delete_reports_by_batch(region, detect_time)
        return {"ok": True, "deleted": count}
    else:
        raise HTTPException(400, "Must provide ids or region+detect_time")


@router.get("/reports/batch-download")
async def batch_download_reports(
    region: str = Query(...),
    detect_time: str = Query(...),
):
    batch_reports = get_reports_by_batch(region, detect_time)
    if not batch_reports:
        raise HTTPException(404, "No reports found for this batch")

    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for rep in batch_reports:
            full_rep = get_report(rep["id"])
            if not full_rep or not full_rep.get("results"):
                continue

            from models.station import DetectionResult, FlagLevel
            from models.station import Report as ReportModel
            from datetime import datetime

            results = []
            for r in full_rep.get("results", []):
                results.append(DetectionResult(
                    report_id=r.get("report_id", rep["id"]),
                    station_id=r.get("station_id", ""),
                    detector=r.get("detector", ""),
                    data_type=r.get("data_type", ""),
                    datetime=r.get("datetime", ""),
                    value=r.get("value", 0),
                    expected_value=r.get("expected_value"),
                    deviation=r.get("deviation"),
                    trigger_rule=r.get("trigger_rule", ""),
                    flag_level=FlagLevel(r.get("flag_level", "Info")),
                    detail=r.get("detail", ""),
                ))

            report = ReportModel(
                id=rep["id"],
                station_name=rep.get("station_name", ""),
                created_at=datetime.fromisoformat(rep["created_at"]),
                total_flags=rep.get("total_flags", 0),
                severe_count=rep.get("severe_count", 0),
                warning_count=rep.get("warning_count", 0),
                general_count=rep.get("general_count", 0),
                info_count=rep.get("info_count", 0),
            )

            tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
            output_path = tmp.name
            tmp.close()

            export_excel(report, results, output_path)
            filename = f"{rep.get('station_name', 'report')}.xlsx"
            zf.write(output_path, filename)
            try:
                os.unlink(output_path)
            except OSError:
                pass

    zip_buffer.seek(0)
    safe_region = region.replace("/", "_").replace("\\", "_")
    safe_time = detect_time.replace(":", "-").replace("/", "-").replace(" ", "_")
    zip_filename = f"{safe_region}_{safe_time}.zip"

    return StreamingResponse(
        zip_buffer,
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{zip_filename}"'},
    )


@router.get("/reports/batch-results")
async def get_batch_results(
    region: str = Query(...),
    detect_time: str = Query(...),
):
    results = get_batch_detection_results(region, detect_time)
    for r in results:
        r.setdefault("report_id", r.get("report_id", ""))
    return results


@router.get("/reports/export-excel")
async def export_reports_excel(
    view_type: str = Query("detail"),
    regions: str = Query(None),
    stations: str = Query(None),
    start_time: str = Query(None),
    end_time: str = Query(None),
    risk_filter: str = Query(None),
    batch_detect_time: str = Query(None),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    region_list = [r.strip() for r in regions.split(",") if r.strip()] if regions else None
    station_list = [s.strip() for s in stations.split(",") if s.strip()] if stations else None

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    if view_type == "summary":
        ws = wb.active
        ws.title = "批次汇总"
        columns = ["所属区域", "检测批次时间", "站点总数", "严重", "警告", "一般", "提示", "总问题条数"]
        ws.append(columns)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        page = 1
        while True:
            batches, total = query_reports_grouped(
                regions=region_list, stations=station_list,
                start_time=start_time, end_time=end_time,
                risk_filter=risk_filter,
                batch_detect_time=batch_detect_time,
                page=page, page_size=1000,
            )
            if not batches:
                break
            for b in batches:
                ws.append([
                    b["region_name"],
                    b["created_at"],
                    b["station_count"],
                    b["severe_count"],
                    b["warning_count"],
                    b["general_count"],
                    b["info_count"],
                    b["total_flags"],
                ])
            if len(batches) < 1000:
                break
            page += 1
    else:
        ws = wb.active
        ws.title = "站点明细"
        columns = ["站点", "所属区域", "检测时间", "严重", "警告", "一般", "提示", "总计"]
        ws.append(columns)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        page = 1
        while True:
            records, total = query_reports(
                regions=region_list, stations=station_list,
                start_time=start_time, end_time=end_time,
                risk_filter=risk_filter,
                batch_detect_time=batch_detect_time,
                page=page, page_size=1000,
            )
            if not records:
                break
            for r in records:
                ws.append([
                    r["station_name"],
                    r.get("region_name", ""),
                    r["created_at"],
                    r["severe_count"],
                    r["warning_count"],
                    r["general_count"],
                    r["info_count"],
                    r["total_flags"],
                ])
            if len(records) < 1000:
                break
            page += 1

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    output_path = tmp.name
    tmp.close()

    wb.save(output_path)

    filename = f"检测报告_{view_type}_{start_time or 'all'}_{end_time or 'all'}.xlsx"
    return FileResponse(
        output_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=filename,
    )


@router.get("/report/{report_id}/summary")
async def get_report_summary(report_id: str):
    rep = get_report(report_id)
    if not rep:
        raise HTTPException(404, "Report not found")

    detector_stats = {}
    for r in rep.get("results", []):
        det = r["detector"]
        lvl = r["flag_level"]
        detector_stats.setdefault(det, {"Severe": 0, "Warning": 0, "General": 0, "Info": 0})
        detector_stats[det][lvl] += 1

    return {
        "report_id": report_id,
        "station_name": rep.get("station_name"),
        "created_at": rep.get("created_at"),
        "total_flags": rep.get("total_flags"),
        "severe_count": rep.get("severe_count"),
        "warning_count": rep.get("warning_count"),
        "general_count": rep.get("general_count"),
        "info_count": rep.get("info_count"),
        "detector_stats": detector_stats,
    }


@router.get("/export/{report_id}")
async def export_report(report_id: str):
    rep = get_report(report_id)
    if not rep:
        raise HTTPException(404, "Report not found")

    from models.station import DetectionResult, FlagLevel
    from models.station import Report as ReportModel
    from datetime import datetime

    results = []
    for r in rep.get("results", []):
        results.append(DetectionResult(
            report_id=r.get("report_id", report_id),
            station_id=r.get("station_id", ""),
            detector=r.get("detector", ""),
            data_type=r.get("data_type", ""),
            datetime=r.get("datetime", ""),
            value=r.get("value", 0),
            expected_value=r.get("expected_value"),
            deviation=r.get("deviation"),
            trigger_rule=r.get("trigger_rule", ""),
            flag_level=FlagLevel(r.get("flag_level", "Info")),
            detail=r.get("detail", ""),
        ))

    report = ReportModel(
        id=rep["id"],
        station_name=rep.get("station_name", ""),
        created_at=datetime.fromisoformat(rep["created_at"]),
        total_flags=rep.get("total_flags", 0),
        severe_count=rep.get("severe_count", 0),
        warning_count=rep.get("warning_count", 0),
        general_count=rep.get("general_count", 0),
        info_count=rep.get("info_count", 0),
    )

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    output_path = tmp.name
    tmp.close()

    export_excel(report, results, output_path)

    filename = f"{rep.get('station_name', 'report')}.xlsx"
    return FileResponse(
        output_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=filename,
    )
